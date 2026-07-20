import re
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


HEADER_LABEL = "Element Name (idShort)"
VALUE_LABEL = "Actual Value"
EXAMPLE_LABEL = "Example Value"
OBLIGATION_LABEL = "Obligation"
FIELD_TYPE_LABEL = "Data Style / Field Type"


@dataclass(frozen=True)
class FieldRecord:
    sheet_name: str
    table_name: str
    row_number: int
    id_short: str
    obligation: str
    field_type: str | None
    example_value: Any
    actual_value: Any

    @property
    def mandatory(self) -> bool:
        return self.obligation.lower().startswith("mandatory")

    @property
    def optional(self) -> bool:
        return self.obligation.lower().startswith("optional")

    @property
    def has_actual_value(self) -> bool:
        return not is_empty(self.actual_value)


@dataclass(frozen=True)
class TableRecord:
    sheet_name: str
    title: str
    name: str
    header_row: int
    fields: tuple[FieldRecord, ...]

    def field(self, id_short: str) -> FieldRecord | None:
        normalized = normalize_name(id_short)
        return next(
            (
                field
                for field in self.fields
                if normalize_name(field.id_short) == normalized
            ),
            None,
        )


@dataclass(frozen=True)
class SheetData:
    name: str
    tables: tuple[TableRecord, ...]

    @property
    def root_table(self) -> TableRecord:
        if not self.tables:
            raise ValueError(f"Worksheet {self.name!r} contains no data tables")
        return self.tables[0]

    def named_tables(self, name: str) -> tuple[TableRecord, ...]:
        normalized = normalize_name(name)
        return tuple(
            table for table in self.tables if normalize_name(table.name) == normalized
        )


@dataclass(frozen=True)
class MissingField:
    sheet_name: str
    table_name: str
    row_number: int
    id_short: str
    reason: str

    def __str__(self) -> str:
        return (
            f"{self.sheet_name} -> {self.table_name} -> {self.id_short} "
            f"(row {self.row_number}: {self.reason})"
        )


@dataclass(frozen=True)
class WorkbookData:
    sheets: dict[str, SheetData]


def read_workbook(path: Path, sheet_names: Iterable[str]) -> WorkbookData:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets: dict[str, SheetData] = {}

    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            sheets[sheet_name] = extract_sheet_data(workbook[sheet_name])

    workbook.close()
    return WorkbookData(sheets=sheets)


def extract_sheet_data(sheet: Worksheet) -> SheetData:
    rows = list(sheet.iter_rows(values_only=True))
    tables: list[TableRecord] = []
    row_index = 0

    while row_index < len(rows):
        row = rows[row_index]
        key_index = _column_index(row, HEADER_LABEL, contains=True)
        if key_index is None:
            row_index += 1
            continue

        title = _table_title(rows, row_index)
        table_name = _table_name(title, sheet.title)
        value_index = _column_index(row, VALUE_LABEL)
        example_index = _column_index(row, EXAMPLE_LABEL)
        obligation_index = _column_index(row, OBLIGATION_LABEL)
        field_type_index = _column_index(row, FIELD_TYPE_LABEL)
        header_row = row_index + 1
        row_index += 1
        fields: list[FieldRecord] = []

        while row_index < len(rows):
            data_row = rows[row_index]
            if key_index >= len(data_row) or is_empty(data_row[key_index]):
                break

            fields.append(
                FieldRecord(
                    sheet_name=sheet.title,
                    table_name=table_name,
                    row_number=row_index + 1,
                    id_short=str(data_row[key_index]).strip(),
                    obligation=str(_cell(data_row, obligation_index) or "").strip(),
                    field_type=_text_or_none(_cell(data_row, field_type_index)),
                    example_value=_clean_value(_cell(data_row, example_index)),
                    actual_value=_clean_value(_cell(data_row, value_index)),
                )
            )
            row_index += 1

        tables.append(
            TableRecord(
                sheet_name=sheet.title,
                title=title,
                name=table_name,
                header_row=header_row,
                fields=tuple(fields),
            )
        )

    return SheetData(name=sheet.title, tables=tuple(tables))


def is_empty(value: Any) -> bool:
    return value is None or str(value).strip().lower() in {"", "n/a"}


def normalize_name(value: str) -> str:
    normalized = re.sub(r"__\d+__$", "", value.strip())
    if normalized.lower() == "documnet":
        return "document"
    return normalized.lower()


def _table_title(rows: list[tuple[Any, ...]], header_index: int) -> str:
    for index in range(header_index - 1, -1, -1):
        if rows[index] and not is_empty(rows[index][0]):
            return str(rows[index][0]).strip()
    return f"Table at row {header_index + 1}"


def _table_name(title: str, sheet_name: str) -> str:
    quoted = re.findall(r"['‘’]([^'‘’]+)['‘’]", title)
    if quoted:
        return quoted[-1].strip()

    scoped = re.search(
        r"\b(?:SMC|SML|Submodel)\s*:\s*([^()]+?)(?:\)|$)", title, re.I
    )
    if scoped:
        return scoped.group(1).strip()

    return sheet_name


def _column_index(
    row: tuple[Any, ...], label: str, *, contains: bool = False
) -> int | None:
    for index, cell in enumerate(row):
        if cell is None:
            continue
        text = str(cell).strip()
        if (contains and label in text) or (not contains and text == label):
            return index
    return None


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    return row[index] if index is not None and index < len(row) else None


def _clean_value(value: Any) -> Any:
    return value.strip() if isinstance(value, str) else value


def _text_or_none(value: Any) -> str | None:
    return None if is_empty(value) else str(value).strip()
